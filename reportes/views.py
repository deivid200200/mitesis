from django.http import HttpResponse
from django.views.generic import TemplateView
from django.contrib.auth.mixins import LoginRequiredMixin
from django.template.loader import render_to_string
from django.contrib import messages
from django.shortcuts import redirect
from datetime import datetime
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side

from academico.models import Expediente, ActaCulminacion, Calificacion, InscripcionModulo


class GenerarExpedientePDFView(LoginRequiredMixin, TemplateView):
    """Generar expediente en PDF del estudiante"""
    template_name = 'reportes/expediente_pdf.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        estudiante_id = self.kwargs.get('estudiante_id')
        
        try:
            expediente = Expediente.objects.get(estudiante_id=estudiante_id)
            calificaciones = Calificacion.objects.filter(estudiante=expediente.estudiante)
            inscripciones = InscripcionModulo.objects.filter(estudiante=expediente.estudiante)
            
            context.update({
                'expediente': expediente,
                'calificaciones': calificaciones,
                'inscripciones': inscripciones,
                'fecha_generacion': datetime.now(),
            })
        except Expediente.DoesNotExist:
            context['error'] = "No se encontró expediente"
        
        return context
    
    def render_to_response(self, context, **response_kwargs):
        try:
            from weasyprint import HTML, CSS
        except ImportError:
            messages.error(self.request, "La generación de PDFs no está disponible. Instale WeasyPrint con sus dependencias del sistema.")
            return redirect('dashboard')
            
        if 'error' in context:
            messages.error(self.request, context['error'])
            return redirect('dashboard')
        
        html_string = render_to_string(self.template_name, context)
        html = HTML(string=html_string, base_url=self.request.build_absolute_uri('/'))
        
        pdf_file = html.write_pdf()
        
        response = HttpResponse(pdf_file, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="expediente_{context["expediente"].estudiante.matricula}.pdf"'
        return response


class GenerarActaCulminacionPDFView(LoginRequiredMixin, TemplateView):
    """Generar acta de culminación en PDF"""
    template_name = 'reportes/acta_culminacion_pdf.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        acta_id = self.kwargs.get('acta_id')
        
        try:
            acta = ActaCulminacion.objects.get(id=acta_id)
            expediente = acta.expediente
            calificaciones = Calificacion.objects.filter(estudiante=expediente.estudiante)
            
            context.update({
                'acta': acta,
                'expediente': expediente,
                'calificaciones': calificaciones,
                'fecha_generacion': datetime.now(),
            })
        except ActaCulminacion.DoesNotExist:
            context['error'] = "No se encontró acta"
        
        return context
    
    def render_to_response(self, context, **response_kwargs):
        try:
            from weasyprint import HTML, CSS
        except ImportError:
            messages.error(self.request, "La generación de PDFs no está disponible. Instale WeasyPrint con sus dependencias del sistema.")
            return redirect('dashboard')
            
        if 'error' in context:
            messages.error(self.request, context['error'])
            return redirect('dashboard')
        
        html_string = render_to_string(self.template_name, context)
        html = HTML(string=html_string, base_url=self.request.build_absolute_uri('/'))
        
        pdf_file = html.write_pdf()
        
        response = HttpResponse(pdf_file, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="acta_culminacion_{context["acta"].numero_acta}.pdf"'
        return response


class GenerarActaNotasExcelView(LoginRequiredMixin, TemplateView):
    """Generar acta de notas en Excel"""
    
    def get(self, request, *args, **kwargs):
        estudiante_id = kwargs.get('estudiante_id')
        
        try:
            expediente = Expediente.objects.get(estudiante_id=estudiante_id)
            calificaciones = Calificacion.objects.filter(estudiante=expediente.estudiante)
            
            # Crear workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Notas"
            
            # Estilos
            header_font = Font(bold=True, size=12)
            header_alignment = Alignment(horizontal='center', vertical='center')
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Encabezado
            ws['A1'] = "ACTA DE CALIFICACIONES"
            ws['A1'].font = Font(bold=True, size=14)
            ws.merge_cells('A1:F1')
            
            ws['A2'] = f"Estudiante: {expediente.estudiante.usuario.get_full_name()}"
            ws['A3'] = f"Matrícula: {expediente.estudiante.matricula}"
            ws['A4'] = f"Maestría: {expediente.maestria.nombre}"
            ws['A5'] = f"Promedio General: {expediente.promedio_general}"
            
            # Encabezados de tabla
            row = 7
            headers = ['Asignatura', 'Código', 'Créditos', 'Parcial 1', 'Parcial 2', 'Trabajo', 'Nota Final', 'Estado']
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col)
                cell.value = header
                cell.font = header_font
                cell.alignment = header_alignment
                cell.border = border
            
            # Datos
            row = 8
            for cal in calificaciones:
                ws.cell(row=row, column=1).value = cal.asignatura.nombre
                ws.cell(row=row, column=2).value = cal.asignatura.codigo
                ws.cell(row=row, column=3).value = cal.asignatura.creditos
                ws.cell(row=row, column=4).value = cal.nota_parcial1 or '-'
                ws.cell(row=row, column=5).value = cal.nota_parcial2 or '-'
                ws.cell(row=row, column=6).value = cal.nota_trabajo or '-'
                ws.cell(row=row, column=7).value = cal.nota_final
                ws.cell(row=row, column=8).value = cal.get_estado_display()
                
                for col in range(1, 9):
                    ws.cell(row=row, column=col).border = border
                
                row += 1
            
            # Ancho de columnas
            ws.column_dimensions['A'].width = 30
            ws.column_dimensions['B'].width = 15
            ws.column_dimensions['C'].width = 10
            ws.column_dimensions['D'].width = 12
            ws.column_dimensions['E'].width = 12
            ws.column_dimensions['F'].width = 12
            ws.column_dimensions['G'].width = 12
            ws.column_dimensions['H'].width = 12
            
            # Guardar a BytesIO
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            
            response = HttpResponse(
                output.getvalue(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="notas_{expediente.estudiante.matricula}.xlsx"'
            return response
            
        except Expediente.DoesNotExist:
            messages.error(request, "No se encontró expediente")
            return redirect('dashboard')


def generar_numero_acta():
    """Generar número único para el acta de culminación"""
    from django.utils import timezone
    now = timezone.now()
    count = ActaCulminacion.objects.count() + 1
    return f"ACTA-{now.year}{now.month:02d}-{count:05d}"


def crear_acta_culminacion(estudiante, maestria):
    """Crear acta de culminación cuando se completan todas las asignaturas"""
    expediente, created = Expediente.objects.get_or_create(
        estudiante=estudiante,
        maestria=maestria
    )
    
    # Verificar que todas las asignaturas estén aprobadas
    calificaciones = Calificacion.objects.filter(
        estudiante=estudiante,
        asignatura__modulo__maestria=maestria
    )
    
    todas_aprobadas = all(
        cal.estado == 'aprobada' for cal in calificaciones
    )
    
    if todas_aprobadas and calificaciones.exists():
        acta, created = ActaCulminacion.objects.get_or_create(
            expediente=expediente,
            defaults={
                'numero_acta': generar_numero_acta(),
                'promedio_final': expediente.promedio_general,
                'titulo_otorgado': f"Magister en {maestria.nombre}",
            }
        )
        return acta
    
    return None


